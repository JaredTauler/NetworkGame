import json

import pygame
import pygame.draw
import pygame.gfxdraw as gfx
import client
import game
import server
import math
from timer import Timer
t = Timer()
import pytmx
import xml.etree.ElementTree as ET

import pymunk as pm
from openpyxl import load_workbook

from function import *

class Player(pg.sprite.Sprite):

	def __init__(self, screen, space, netplayer):
		pg.sprite.Sprite.__init__(self)
		# net
		self.netplayer = netplayer
		self.ticklist = []

		# pm
		self.body = pm.Body()
		self.body.position = (100,100)
		self.shape = pm.Poly.create_box(self.body, (30,30))
		self.shape.density = 10
		self.shape.elasticity = 1
		space.add(self.body, self.shape)
		self.shape.collision_type = 1

		# pg
		self.surf = pg.Surface((20, 20))
		self.rect = self.surf.get_rect()
		self.rect.center = screen.rect.center

		pg.draw.rect(self.surf, (255, 0, 0), self.surf.get_rect())

	def getsurf(self):
		return self.surf

	def update(self, screen, group, input, space):

		while len(self.ticklist) != 0:
			t = self.ticklist[0]
			self.body.position = t["location"]
			self.body.velocity = t["velocity"]
			self.ticklist.pop(0)
			print("BRUH")

		key = {"left": 97, "up": 119, "down": 115, "right": 100}
		dir = [0,0]
		if not self.netplayer:
			if pg.key.get_pressed()[key["left"]]:
				dir = SumTup(dir, (-1, 0))
			if pg.key.get_pressed()[key["right"]]:
				dir = SumTup(dir, (1, 0))
			if pg.key.get_pressed()[key["up"]]:
				dir = SumTup(dir, (0, -1))
			if pg.key.get_pressed()[key["down"]]:
				dir = SumTup(dir, (0, 1))
		mvspd = .2
		self.body.velocity = SumTup((dir[0]*mvspd, dir[1]*mvspd), self.body.velocity)

		if not self.netplayer:
			# Update camera location
			screen.location = (
				-self.body.position.x + screen.rect.center[0],
				-self.body.position.y + screen.rect.center[1]
			)

class Tile(pg.sprite.Sprite):
	def __init__(self, size, loc, space, textureid, tiles):
		pg.sprite.Sprite.__init__(self)
		self.textureid = textureid # Remember texture ID rather than texture

		# pm
		self.body = pm.Body(body_type=pm.Body.STATIC)
		self.body.position = loc
		self.shape = pm.Poly.create_box(self.body, size)
		self.shape.density = 1
		self.shape.elasticity = 1
		space.add(self.body, self.shape)
		self.shape.collision_type = 1

		self.tiles = tiles


	def update(self, screen, group, input, space):
		pass

	def getsurf(self):
		self.surf = pg.Surface((20, 20))
		pg.draw.rect(self.surf, (255, 0, 0), self.surf.get_rect())
		return self.surf
		# return self.tiles[self.textureid]

class Game:
	def __init__(self, screen, Forclient):
		self.screen = screen

		if not Forclient:
			pg.display.set_caption(str("SERVER"))
			self.server = server.Server()
		else:
			pg.display.set_caption(str("CLIENT"))
		self.net = client.Network()
		import socket
		self.net.connect((socket.gethostname(), 5059))

		self.lastresponse = None

		self.space = pm.Space()  # PyMunk simulation
		self.space.gravity = (0, .1)
		self.tiles = {}
		self.group = {}
		self.group["entity"] = []

		self.group["world"] = []

		self.LoadMap()

		self.group["player"] = []
		self.group["player"].append(Player(screen, self.space, False))

	# def LoadMap(self):
	# 	# self.group["world"].append(Tile((40, 40), (0,0), self.space))
	# 	loc = ("map.xlsx")
	#
	# 	wb = load_workbook(loc)
	# 	ws = wb.active
	# 	m = 70
	# 	for row in range(0, ws.max_row):
	# 		for i, col in enumerate(ws.iter_cols(1, ws.max_column)):
	# 			if col[row].value == 1:
	# 				self.group["world"].append(Tile((m, m), (i*m, row*m), self.space))
	def LoadMap(self):
		dir = "map/map1/"  # base dir from which files being getten from.

		# PARSE TSX FILE
		tree = ET.parse(dir + "base.tsx")
		root = tree.getroot()
		self.tiles = {} # surfs
		for child in root: # For elem in XML
			if child.tag == "tile": # Make sure getting a tile
				for i in child:
					loc = dir + i.attrib["source"]
					# Dont care about any other attributes, pygame can automatically determine what size surf should be.
					surf = pygame.image.load(dir + i.attrib["source"])
					self.tiles[int(child.attrib["id"]) + 1] = surf
					print(i.attrib["source"],int(child.attrib["id"]) + 1)

		# PARSE MAP
		# tree = ET.parse(dir + ".tmx")
		# root = tree.getroot()
		# import csv
		# for i in csv.reader(root[1][0].text):
		# 	print(i)
		# # print(root[1][0].text)

		tm = pytmx.load_pygame(dir + ".tmx") # FIXME stop this from auto loading images... ill do it myself thank you.
		for x, y, gid, in tm.get_layer_by_name("terrain", ):
			# print(gid)
			if gid == 0:
				continue
			n=20
			self.group["world"].append(Tile((n, n), (x*n, y*n), self.space, gid, self.tiles))


	def update(self, screen, group, input):
		if self.net.response != []: print(self.net.response)
		# Made in such a way that new ticks can come in while this process is happening.
		while self.net.response:
			for res in self.net.response[0]: # First response in list of responses
				for client_id in res: # Process each computers info seperate
					# Finally processing ticks.
					for tick in res[client_id]:
						print(tick)
						if type(tick) is str:
							tick = json.loads(tick)
						if client_id == "0":
							if res.get("id"):
								id = res.get("id")
								self.client_id = id
								print("Client ID set to " + str(id))
						else:
							for player in tick.get("netplayer"): # If client has more than 1 player
								print("Moving Netplayer")
								# Create netplayer
								if not self.group.get(client_id):
									self.group[client_id] = []
								if len(self.group.get(client_id)) == 0:
									self.group[client_id].append(Player(screen, True, self.space))
								# Pass tickdata to said netplayer for processing
								self.group[client_id][0].ticklist.append(
									tick.get("netplayer")[player]
								)

			self.net.response.pop(0)  # Tick has been processed, remove it from the list

		# Step pymunk sim
		self.space.step(1)

		screen.surf.fill([121, 100, 100])
		# Update entities
		t.start()
		for e in self.group.values():
			for tick in e:
				act = tick.update(screen, self.group, input, self.space)
				# draw_poly(tick, screen)
				# try:
				# pos = tick.shape.body.position
				loc = SumTup(tick.shape.body.position, screen.location)
				# print(screen.location[1], loc[1])
				if (
					screen.location[0] < loc[0] + screen.location[0] # Left
					and screen.location[0] + screen.rect.x < loc[0] - screen.location[0]
				):
					screen.surf.blit(tick.getsurf(), loc)
				# shape = tick.shape
				# verts = []
				# for v in shape.get_vertices():
				# 	n = SumTup(
				# 			v.rotated(shape.body.angle),
				# 			screen.location,
				# 			shape.body.position
				# 	)
				# # 	# n = SumTup(
				# # 	# 		v.rotated(shape.body.angle),
				# # 	# )
				# 	verts.append(n)
				# # print(shape.body.position)
				#
				# # try:
				# # 	print(shape)
				# # 	print(shape.body.position)
				# try:
				# 	pg.draw.polygon(screen.surf, [255,255,255], verts)
				# 	# gfx.textured_polygon(screen.surf, verts, tick.getsurf(), int(shape.body.position[0]), int(shape.body.position[1]))
				# except: pass
		t.stop()


		# Send data to server
		data = {}
		for i, p in enumerate(self.group["player"]):
			pdata = {"location": p.body.position, "velocity": p.body.velocity}
			data["netplayer"] = {i: pdata}

		# Dont send if already sent same data last time.
		if self.lastresponse != data:
			self.net.send(data)
		self.lastresponse = data
